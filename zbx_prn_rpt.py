import requests
import re
import openpyxl
import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from requests.packages.urllib3.exceptions import InsecureRequestWarning

# Отключение предупреждений о небезопасных запросах
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Конфигурация Zabbix
ZABBIX_URL = "https://<АДРЕС_ZABBIX_Сервера>.ru/api_jsonrpc.php"
USERNAME = "Admin"
PASSWORD = "PASSWORD"
HEADERS = {"Content-Type": "application/json-rpc"}

# Шаблон для имен принтеров
NAME_PATTERN = re.compile(r"^\d{4}-P\d{2}$")

# Параметры для разных производителей
PRINTER_PARAMS = {
    "Kyocera": {
        "group": "_Printers",
        "keys": {
            "model": "model.device",
            "pages": "a4.device",
            "sn": "sn.device"
        }
    },
    "Brother": {
        "group": "_Printers_Brother",
        "keys": {
            "pages": "mib-2.43.10.2.1.4.1.1",
            "sn": "mib-2.43.5.1.1.17.1"
        }
    }
}

def zabbix_request(method, params, auth_token=None):
    """Выполнение запроса к Zabbix API"""
    payload = {
        "jsonrpc": "2.0",
        "method": method,
        "params": params,
        "id": 1,
        "auth": auth_token
    }
    try:
        response = requests.post(
            ZABBIX_URL,
            json=payload,
            headers=HEADERS,
            verify=False,
            timeout=30
        )
        response.raise_for_status()
        return response.json().get("result")
    except Exception as e:
        print(f"Ошибка запроса: {str(e)}")
        return None

def get_auth_token():
    """Аутентификация в Zabbix и получение токена"""
    params = {
        "user": USERNAME,
        "password": PASSWORD
    }
    result = zabbix_request("user.login", params)
    if not result:
        raise Exception("Ошибка аутентификации в Zabbix")
    return result

def get_host_group(auth_token, group_name):
    """Получение ID группы по имени"""
    params = {
        "filter": {"name": group_name}
    }
    groups = zabbix_request("hostgroup.get", params, auth_token)
    return groups[0]["groupid"] if groups else None

def get_hosts(auth_token, group_id):
    """Получение хостов из группы с интерфейсами"""
    params = {
        "groupids": group_id,
        "output": ["hostid", "host", "name"],
        "selectInterfaces": ["ip"],
        "filter": {"status": "0"}  # только активные хосты
    }
    return zabbix_request("host.get", params, auth_token) or []

def get_items(auth_token, hostids, keys):
    """Получение значений элементов для хостов по точным ключам"""
    if not hostids or not keys:
        return []
    
    params = {
        "hostids": hostids,
        "output": ["lastvalue", "key_", "hostid"],
        "filter": {"key_": list(keys.values())}
    }
    return zabbix_request("item.get", params, auth_token) or []

def process_printers():
    """Основная логика сбора данных"""
    # Аутентификация
    auth_token = get_auth_token()
    if not auth_token:
        raise Exception("Ошибка аутентификации в Zabbix")
    print(f"Успешная аутентификация, токен: {auth_token[:10]}...")

    # Сбор данных по принтерам
    all_printers = []
    
    for printer_type, params in PRINTER_PARAMS.items():
        print(f"\nОбработка группы: {params['group']} ({printer_type})")
        
        # Получение ID группы
        group_id = get_host_group(auth_token, params["group"])
        if not group_id:
            print(f"Группа {params['group']} не найдена!")
            continue
        print(f"ID группы: {group_id}")

        # Получение хостов группы
        hosts = get_hosts(auth_token, group_id)
        if not hosts:
            print(f"Хосты в группе {params['group']} не найдены!")
            continue
            
        print(f"Найдено хостов в группе: {len(hosts)}")
        
        # Фильтрация по шаблону имени (поле host)
        filtered_hosts = [
            h for h in hosts 
            if NAME_PATTERN.match(h["host"])
        ]
        
        print(f"Хостов после фильтрации: {len(filtered_hosts)}")
        
        if not filtered_hosts:
            print("Нет хостов, соответствующих шаблону 0000-P00")
            continue

        # Получаем элементы по точным ключам
        hostids = [h["hostid"] for h in filtered_hosts]
        items = get_items(auth_token, hostids, params["keys"])
        
        if not items:
            print(f"Элементы данных не найдены для ключей: {list(params['keys'].values())}")
        else:
            print(f"Получено элементов данных: {len(items)}")
        
        # Сопоставляем данные с хостами
        for host in filtered_hosts:
            printer_data = {
                "host": host["host"],
                "name": host["name"],
                "ip": host["interfaces"][0]["ip"] if host["interfaces"] else "N/A",
                "type": printer_type
            }
            
            # Для Brother устанавливаем модель по умолчанию
            if printer_type == "Brother":
                printer_data["model"] = "Brother"
            
            # Ищем соответствующие элементы
            host_items = [i for i in items if i["hostid"] == host["hostid"]] if items else []
            
            for key_name, key_value in params["keys"].items():
                # Находим элемент по точному соответствию ключа
                item_value = next(
                    (i["lastvalue"] for i in host_items if i["key_"] == key_value), 
                    "N/A"
                )
                printer_data[key_name] = item_value
            
            all_printers.append(printer_data)
    
    # Сортировка по имени хоста
    return sorted(all_printers, key=lambda x: x["host"])

def create_excel_report(printers_data):
    """Создание Excel-отчета с датой в имени файла"""
    # Формируем имя файла с текущей датой
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"printers_report_{today}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по принтерам"
    
    # Добавляем дату формирования отчета в заголовок
    report_date = datetime.datetime.now().strftime("%d.%m.%Y")
    ws.cell(row=1, column=1, value=f"Отчет по счетчикам принтеров на {report_date}")
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Заголовки таблицы
    headers = [
        "Техническое имя", 
        "Отображаемое имя",
        "IP адрес", 
        "Производитель", 
        "Модель", 
        "Серийный номер", 
        "Всего страниц"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Данные
    for row, printer in enumerate(printers_data, 3):  # Начинаем с 3 строки
        ws.cell(row=row, column=1, value=printer["host"])
        ws.cell(row=row, column=2, value=printer.get("name", "N/A"))
        ws.cell(row=row, column=3, value=printer["ip"])
        ws.cell(row=row, column=4, value=printer["type"])
        ws.cell(row=row, column=5, value=printer.get("model", "N/A"))
        ws.cell(row=row, column=6, value=printer.get("sn", "N/A"))
        ws.cell(row=row, column=7, value=printer.get("pages", "N/A"))
    
    # Авто-ширина колонок с обработкой объединенных ячеек
    for col_idx in range(1, 8):  # У нас 7 колонок (A-G)
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Проверяем только ячейки с данными (начиная со строки 2)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            
            # Пропускаем объединенные ячейки
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
                
            try:
                value = str(cell.value) if cell.value else ""
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
                
        # Устанавливаем ширину с ограничением
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename

def main():
    """Основная функция"""
    try:
        print("Сбор данных о принтерах...")
        printers_data = process_printers()
        
        if not printers_data:
            print("Не найдено принтеров, соответствующих критериям")
            return
        
        print(f"\nНайдено принтеров для отчета: {len(printers_data)}")
        filename = create_excel_report(printers_data)
        print(f"Отчет успешно сохранен: {filename}")
        
    except Exception as e:
        print(f"Критическая ошибка: {str(e)}")

if __name__ == "__main__":
    main()
